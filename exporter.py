"""
CHIMERA Suite – Daten- und Bericht-Exporter
=============================================
Liest Checkpoint-JSON-Daten (aus run_endurance.py oder run-Befehlen) und
exportiert sie in mehrere Zielformate.

Unterstützte Ausgabeformate:
    json    Aggregierte Rohdaten als JSON
    csv     Tabellarische Übersicht als CSV
    html    Interaktiver HTML-Bericht mit Diagramm (kein JS-Framework)
    md      Markdown-Zusammenfassung
    xlsx    Excel-Tabelle (erfordert openpyxl)

Direkte Verwendung:
    from exporter import ChimeraExporter
    exp = ChimeraExporter(rows, pathlib.Path("./results"))
    exp.export(["json", "csv", "html"], base_name="mein_bericht")
"""

import csv
import json
import pathlib
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Sequence


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

_WORKLOAD_ORDER = [
    "ycsb_a", "ycsb_b", "ycsb_c", "ycsb_d", "ycsb_e", "ycsb_f",
    "vector_search", "graph_traversal",
    "ts_ingest_raw", "ts_ingest_batch", "ts_range_query",
    "geo_intersects", "geo_contains", "geo_range_query",
    "process_discovery_alpha", "process_discovery_heuristic", "process_conformance",
    "llm_serving_latency", "llm_serving_throughput",
    "ml_policy_eval", "ethics_eval_pipeline",
    "lora_load_apply", "lora_switch_overhead",
    "rag_retrieval_quality", "rag_qa_e2e",
    "tpc_c_mix",
]


def _mean(values: List[float]) -> float:
    return sum(values) / len(values)


def _stdev(values: List[float]) -> float:
    if len(values) < 2:
        return 0.0
    mean = _mean(values)
    variance = sum((value - mean) ** 2 for value in values) / (len(values) - 1)
    return variance ** 0.5


def _sort_rows(rows: List[Dict[str, Any]], sort_by: str) -> List[Dict[str, Any]]:
    if sort_by == "time":
        return sorted(rows, key=lambda r: r.get("timestamp", ""))
    if sort_by == "metric":
        return sorted(rows, key=lambda r: r.get("throughput_ops_per_sec", 0), reverse=True)
    if sort_by == "alphabetical":
        return sorted(rows, key=lambda r: r.get("workload_id", ""))
    return rows


def _aggregate_by_workload(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Fasst alle Zeilen pro workload_id zu Mittelwerten zusammen."""
    buckets: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        buckets[r.get("workload_id", "unknown")].append(r)

    agg = []
    for wid in _WORKLOAD_ORDER + sorted(set(buckets) - set(_WORKLOAD_ORDER)):
        if wid not in buckets:
            continue
        grp = buckets[wid]
        nums = {
            "throughput_ops_per_sec",
            "mean_latency_ms", "p50_latency_ms",
            "p95_latency_ms", "p99_latency_ms",
            "error_count", "elapsed_seconds",
        }
        row: Dict[str, Any] = {"workload_id": wid, "samples": len(grp)}
        for key in nums:
            vals = [r[key] for r in grp if key in r]
            if vals:
                row[f"{key}_mean"] = round(_mean(vals), 3)
                row[f"{key}_min"] = round(min(vals), 3)
                row[f"{key}_max"] = round(max(vals), 3)
                if len(vals) > 1:
                    row[f"{key}_stdev"] = round(_stdev(vals), 3)
        agg.append(row)
    return agg


def _extract_report_metadata(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not rows:
        return {}

    first = rows[0]
    
    # Klassische Metadaten (Datenbank-Info)
    metadata = {
        "db_system": first.get("db_system"),
        "database_product": first.get("database_product"),
        "database_version": first.get("database_version"),
        "database_status": first.get("database_status"),
        "database_deployment_mode": first.get("database_deployment_mode"),
        "database_cluster_enabled": first.get("database_cluster_enabled"),
        "database_cluster_role": first.get("database_cluster_role"),
        "database_node_count": first.get("database_node_count"),
        "database_shard_count": first.get("database_shard_count"),
        "database_replication_enabled": first.get("database_replication_enabled"),
        "database_replication_role": first.get("database_replication_role"),
        "database_replication_factor": first.get("database_replication_factor"),
        "federation_enabled": first.get("federation_enabled"),
        "federation_members": first.get("federation_members"),
        "llm_enabled": first.get("llm_enabled"),
        "llm_provider": first.get("llm_provider"),
        "llm_model": first.get("llm_model"),
    }
    
    # Neue Compliance-Metadaten (neu)
    metadata.update({
        "result_schema_id": first.get("result_schema_id"),
        "result_schema_version": first.get("result_schema_version"),
        "standard_id": first.get("standard_id"),
        "dataset_provisioning_seed": first.get("dataset_provisioning_seed"),
        "dataset_provisioning_version": first.get("dataset_provisioning_version"),
    })
    
    return {key: value for key, value in metadata.items() if value not in (None, "")}


# ---------------------------------------------------------------------------
# Haupt-Exporter-Klasse
# ---------------------------------------------------------------------------

class ChimeraExporter:
    """Exportiert CHIMERA-Benchmark-Ergebnisse in mehrere Formate.

    Args:
        rows:        Liste von Ergebnis-Dicts (aus Checkpoint-JSON oder run-Befehl).
        output_dir:  Zielverzeichnis für alle Ausgabedateien.
    """

    def __init__(
        self,
        rows: List[Dict[str, Any]],
        output_dir: pathlib.Path,
    ) -> None:
        self._rows = rows
        self._out = pathlib.Path(output_dir)
        self._out.mkdir(parents=True, exist_ok=True)
        self._generated_at = datetime.now(timezone.utc).isoformat()

    # ------------------------------------------------------------------
    # Öffentliche API
    # ------------------------------------------------------------------

    def export(
        self,
        formats: Sequence[str],
        base_name: str = "chimera_export",
        sort_by: str = "time",
        aggregate: bool = False,
    ) -> List[pathlib.Path]:
        """Exportiert in alle gewünschten Formate.

        Args:
            formats:    Sequenz von Format-Strings (z. B. ``["json", "csv", "html"]``).
            base_name:  Präfix für Ausgabedateinamen.
            sort_by:    Sortierung: ``"time"``, ``"metric"``, ``"alphabetical"``.
            aggregate:  Wenn ``True``, Ergebnisse pro Workload zu Mittelwerten falten.

        Returns:
            Liste der geschriebenen Pfade.
        """
        rows = _sort_rows(self._rows, sort_by)
        agg_rows = _aggregate_by_workload(rows) if aggregate else rows

        written: List[pathlib.Path] = []
        dispatch = {
            "json":  self._write_json,
            "csv":   self._write_csv,
            "html":  self._write_html,
            "md":    self._write_markdown,
            "xlsx":  self._write_xlsx,
        }

        for fmt in formats:
            fmt = fmt.strip().lower()
            fn = dispatch.get(fmt)
            if fn is None:
                import warnings
                warnings.warn(f"Unbekanntes Format '{fmt}' – übersprungen.", stacklevel=2)
                continue
            path = fn(agg_rows if aggregate else rows, base_name)
            if path is not None:
                written.append(path)

        return written

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    def _write_json(
        self,
        rows: List[Dict[str, Any]],
        base_name: str,
    ) -> pathlib.Path:
        path = self._out / f"{base_name}.json"
        payload = {
            "chimera_version": "1.0.0",
            "generated_at": self._generated_at,
            "record_count": len(rows),
            "metadata": _extract_report_metadata(rows),
            "results": rows,
        }
        path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
        return path

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    def _write_csv(
        self,
        rows: List[Dict[str, Any]],
        base_name: str,
    ) -> Optional[pathlib.Path]:
        if not rows:
            return None
        path = self._out / f"{base_name}.csv"

        # Alle Spalten ermitteln (Reihenfolge: bekannte zuerst)
        priority = [
            "workload_id", "timestamp", "samples",
            # Standard-Compliance-Felder (neu)
            "result_schema_id", "result_schema_version",
            "standard_id", "standard_required_metrics", "standard_present_metrics",
            "standard_missing_metrics", "standard_metric_coverage_pct",
            # MLPerf-Felder (neu)
            "mlperf_scenario", "mlperf_dataset", "mlperf_quality_target",
            "mlperf_quality_score", "quality_target_passed",
            "mlperf_performance_target_passed", "mlperf_compliance_passed",
            # BEIR-Felder (neu)
            "beir_dataset", "beir_k", "ndcg_at_10", "mrr", "map",
            # RAGAS-Felder (neu)
            "faithfulness", "answer_relevance", "context_precision",
            "context_recall", "ragas_score", "ragbench_score", "ragas_eval_samples",
            # Klassische Metriken
            "benchmark_domain", "quality_score", "stability_score",
            "throughput_ops_per_sec", "throughput_ops_per_sec_mean",
            "mean_latency_ms", "mean_latency_ms_mean",
            "p50_latency_ms", "p50_latency_ms_mean",
            "p95_latency_ms", "p95_latency_ms_mean",
            "p99_latency_ms", "p99_latency_ms_mean",
            "error_count", "error_count_mean",
            "elapsed_seconds", "elapsed_seconds_mean",
            "ingest_rate_points_per_sec", "geo_query_latency_p95_ms",
            "conformance_rate", "tokens_or_inferences_per_sec",
            "retrieval_quality_score",
        ]
        all_keys: List[str] = list(dict.fromkeys(
            [k for k in priority if any(k in r for r in rows)]
            + [k for r in rows for k in r if k not in priority]
        ))

        with open(path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=all_keys, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        return path

    # ------------------------------------------------------------------
    # HTML
    # ------------------------------------------------------------------

    def _write_html(
        self,
        rows: List[Dict[str, Any]],
        base_name: str,
    ) -> pathlib.Path:
        path = self._out / f"{base_name}.html"
        path.write_text(self._render_html(rows, base_name), encoding="utf-8")
        return path

    def _render_html(self, rows: List[Dict[str, Any]], title: str) -> str:
        agg = _aggregate_by_workload(rows)
        metadata = _extract_report_metadata(rows)
        metadata_rows_html = ""
        for key, value in metadata.items():
            metadata_rows_html += f"<tr><td>{key}</td><td>{value}</td></tr>\n"

        # Chart-Daten (Throughput-Balken)
        chart_labels = json.dumps([r["workload_id"] for r in agg])
        chart_throughput = json.dumps([
            r.get("throughput_ops_per_sec_mean", r.get("throughput_ops_per_sec", 0))
            for r in agg
        ])
        chart_p99 = json.dumps([
            r.get("p99_latency_ms_mean", r.get("p99_latency_ms", 0))
            for r in agg
        ])

        # Zeilen der Zusammenfassungs-Tabelle
        agg_rows_html = ""
        for r in agg:
            tput = r.get("throughput_ops_per_sec_mean", r.get("throughput_ops_per_sec", "–"))
            p50  = r.get("p50_latency_ms_mean",  r.get("p50_latency_ms", "–"))
            p95  = r.get("p95_latency_ms_mean",  r.get("p95_latency_ms", "–"))
            p99  = r.get("p99_latency_ms_mean",  r.get("p99_latency_ms", "–"))
            errs = r.get("error_count_mean",      r.get("error_count", 0))
            n    = r.get("samples", 1)
            agg_rows_html += (
                f"<tr>"
                f"<td>{r['workload_id']}</td>"
                f"<td>{_fmt(tput)}</td>"
                f"<td>{_fmt(p50)}</td>"
                f"<td>{_fmt(p95)}</td>"
                f"<td>{_fmt(p99)}</td>"
                f"<td>{_fmt(errs)}</td>"
                f"<td>{n}</td>"
                f"</tr>\n"
            )

        # Rohdaten-Tabelle (alle Zeilen)
        if rows:
            raw_keys = list(dict.fromkeys(k for r in rows for k in r))
            raw_header = "".join(f"<th>{k}</th>" for k in raw_keys)
            raw_body = ""
            for r in rows:
                raw_body += "<tr>" + "".join(
                    f"<td>{r.get(k, '')}</td>" for k in raw_keys
                ) + "</tr>\n"
        else:
            raw_header = ""
            raw_body = ""

        return f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>CHIMERA – {title}</title>
<style>
  :root {{
    --c1:#0173b2;--c2:#de8f05;--c3:#029e73;--c4:#d55e00;
    --c5:#cc78bc;--c6:#ca9161;--c7:#fbafe4;--c8:#949494;
  }}
  body  {{ font-family: system-ui, sans-serif; max-width: 1200px;
            margin: 0 auto; padding: 1.5rem; color: #222; }}
  h1    {{ color: var(--c1); border-bottom: 3px solid var(--c1); padding-bottom:.4rem; }}
  h2    {{ color: #444; margin-top: 2rem; }}
  table {{ border-collapse: collapse; width: 100%; margin: 1rem 0; font-size:.9rem; }}
  th    {{ background: var(--c1); color: #fff; padding: 8px 12px; text-align: left; }}
  td    {{ padding: 6px 12px; border-bottom: 1px solid #ddd; }}
  tr:hover td {{ background: #f0f6ff; }}
  .badge {{ display:inline-block; background:#e8f4e8; color:#1a6b1a;
             border:1px solid #aadaaa; border-radius:4px; padding:2px 8px;
             font-size:.8rem; margin-left:.5rem; }}
  .meta  {{ color:#666; font-size:.85rem; margin-bottom:1rem; }}
  canvas {{ max-height:320px; }}
  .chart-wrap {{ display:grid; grid-template-columns:1fr 1fr; gap:1.5rem;
                 margin:1.5rem 0; }}
  details summary {{ cursor:pointer; font-weight:bold; padding:.4rem 0; }}
  footer {{ margin-top:3rem; border-top:2px solid #ddd; padding-top:1rem;
             color:#888; font-size:.8rem; }}
  @media(max-width:700px){{ .chart-wrap{{grid-template-columns:1fr}} }}
</style>
</head>
<body>

<h1>CHIMERA Benchmark-Bericht
  <span class="badge">Vendor-neutral</span>
</h1>
<p class="meta">Erstellt: {self._generated_at} &nbsp;|&nbsp;
  Datensätze: {len(rows)} &nbsp;|&nbsp; Datei: {title}</p>

<h2>Zusammenfassung</h2>
<table>
  <thead>
    <tr>
      <th>Workload</th>
      <th>Throughput (ops/s)</th>
      <th>p50 ms</th>
      <th>p95 ms</th>
      <th>p99 ms</th>
      <th>Fehler (∅)</th>
                "## Umgebungs- und DB-Metadaten",
                "",
            ]
            if metadata:
                for key, value in metadata.items():
                    lines.append(f"- **{key}**: {value}")
            else:
                lines.append("- Keine Metadaten verfügbar")

      <th>Samples</th>
    </tr>
  </thead>
  <tbody>
    {agg_rows_html}
  </tbody>
</table>

<h2>Umgebungs- und DB-Metadaten</h2>
<table>
    <thead>
        <tr><th>Schlüssel</th><th>Wert</th></tr>
    </thead>
    <tbody>
        {metadata_rows_html or '<tr><td colspan="2">Keine Metadaten verfügbar</td></tr>'}
    </tbody>
</table>

<h2>Diagramme</h2>
<div class="chart-wrap">
  <div><canvas id="chartTput"></canvas></div>
  <div><canvas id="chartP99"></canvas></div>
</div>

<h2>Rohdaten</h2>
<details>
  <summary>Alle {len(rows)} Zeilen anzeigen</summary>
  <div style="overflow-x:auto">
  <table>
    <thead><tr>{raw_header}</tr></thead>
    <tbody>{raw_body}</tbody>
  </table>
  </div>
</details>

<footer>
  <p>CHIMERA Suite – Vendor-neutral Benchmark Framework &nbsp;|&nbsp;
  Farben: Okabe-Ito (barrierefreiheitskonform) &nbsp;|&nbsp;
  Sortierung: alphabetisch &nbsp;|&nbsp; Keine Herstellerlogos</p>
</footer>

<script>
/* Minimales Balken-Diagramm ohne externe Bibliotheken */
(function(){{
  var labels = {chart_labels};
  var colors = ["#0173b2","#de8f05","#029e73","#d55e00",
                "#cc78bc","#ca9161","#fbafe4","#949494"];

  function drawChart(canvasId, data, label, unit) {{
    var canvas = document.getElementById(canvasId);
    if (!canvas) return;
    var ctx = canvas.getContext("2d");
    var W = canvas.offsetWidth || 500;
    var H = 300;
    canvas.width = W; canvas.height = H;
    var pad = {{top:30,right:20,bottom:60,left:65}};
    var dW = W - pad.left - pad.right;
    var dH = H - pad.top - pad.bottom;
    var maxVal = Math.max.apply(null, data) * 1.1 || 1;
    var barW = dW / labels.length * 0.6;
    var gap  = dW / labels.length;

    ctx.clearRect(0,0,W,H);

    /* Achsen */
    ctx.strokeStyle="#ccc"; ctx.lineWidth=1;
    ctx.beginPath();
    ctx.moveTo(pad.left, pad.top);
    ctx.lineTo(pad.left, pad.top+dH);
    ctx.lineTo(pad.left+dW, pad.top+dH);
    ctx.stroke();

    /* Y-Gitter + Beschriftung */
    ctx.fillStyle="#666"; ctx.font="11px system-ui"; ctx.textAlign="right";
    for(var i=0;i<=4;i++){{
      var y = pad.top + dH - (i/4)*dH;
      var v = (i/4)*maxVal;
      ctx.fillText(v.toFixed(v<1?2:1), pad.left-6, y+4);
      ctx.strokeStyle="#eee";
      ctx.beginPath(); ctx.moveTo(pad.left,y); ctx.lineTo(pad.left+dW,y); ctx.stroke();
    }}

    /* Balken */
    for(var j=0;j<data.length;j++){{
      var x = pad.left + gap*j + gap/2 - barW/2;
      var h = (data[j]/maxVal)*dH;
      var y2 = pad.top + dH - h;
      ctx.fillStyle = colors[j % colors.length];
      ctx.fillRect(x, y2, barW, h);

      /* X-Achsenbeschriftung */
      ctx.fillStyle="#444"; ctx.font="10px system-ui";
      ctx.textAlign="center";
      ctx.save();
      ctx.translate(x+barW/2, pad.top+dH+8);
      ctx.rotate(-0.5);
      ctx.fillText(labels[j], 0, 0);
      ctx.restore();
    }}

    /* Titel */
    ctx.fillStyle="#222"; ctx.font="bold 12px system-ui"; ctx.textAlign="center";
    ctx.fillText(label + " (" + unit + ")", W/2, 18);
  }}

  drawChart("chartTput", {chart_throughput}, "Throughput", "ops/s");
  drawChart("chartP99",  {chart_p99},        "P99-Latenz", "ms");
}})();
</script>
</body>
</html>
"""

    # ------------------------------------------------------------------
    # Markdown
    # ------------------------------------------------------------------

    def _write_markdown(
        self,
        rows: List[Dict[str, Any]],
        base_name: str,
    ) -> pathlib.Path:
        path = self._out / f"{base_name}.md"
        agg = _aggregate_by_workload(rows)
        metadata = _extract_report_metadata(rows)

        lines = [
            f"# CHIMERA Benchmark-Bericht – {base_name}",
            "",
            f"> Erstellt: {self._generated_at}  ",
            f"> Datensätze: {len(rows)}",
            "",
            "## Umgebungs- und DB-Metadaten",
            "",
        ]
        if metadata:
            for key, value in metadata.items():
                lines.append(f"- **{key}**: {value}")
        else:
            lines.append("- Keine Metadaten verfügbar")

        # Compliance-Zusammenfassung (neu)
        if rows and rows[0].get("result_schema_id"):
            lines += [
                "",
                "## Standardkonforme Ergebnisse",
                "",
                f"- **Schema-ID**: {rows[0].get('result_schema_id', 'N/A')}",
                f"- **Schema-Version**: {rows[0].get('result_schema_version', 'N/A')}",
            ]
            # Gruppierung nach Standard
            standards = {}
            for row in rows:
                std_id = row.get("standard_id", "unknown")
                if std_id not in standards:
                    standards[std_id] = []
                standards[std_id].append(row)
            
            if standards:
                lines.append("")
                for std_id, std_rows in sorted(standards.items()):
                    lines.append(f"### {std_id}")
                    coverage_vals = [r.get("standard_metric_coverage_pct", 0) for r in std_rows if "standard_metric_coverage_pct" in r]
                    if coverage_vals:
                        avg_coverage = sum(coverage_vals) / len(coverage_vals)
                        lines.append(f"- **Durchschnittliche Metrik-Coverage**: {avg_coverage:.1f}%")

        lines += [
            "",
            "## Leistungszusammenfassung",
            "",
            "| Workload | Throughput (ops/s) | p50 ms | p95 ms | p99 ms | Fehler | Samples |",
            "|---|---:|---:|---:|---:|---:|---:|",
        ]
        for r in agg:
            tput = r.get("throughput_ops_per_sec_mean", r.get("throughput_ops_per_sec", "–"))
            p50  = r.get("p50_latency_ms_mean",  r.get("p50_latency_ms", "–"))
            p95  = r.get("p95_latency_ms_mean",  r.get("p95_latency_ms", "–"))
            p99  = r.get("p99_latency_ms_mean",  r.get("p99_latency_ms", "–"))
            errs = r.get("error_count_mean",      r.get("error_count", 0))
            n    = r.get("samples", 1)
            lines.append(
                f"| {r['workload_id']} | {_fmt(tput)} | {_fmt(p50)} | {_fmt(p95)} | {_fmt(p99)} | {_fmt(errs)} | {n} |"
            )

        lines += [
            "",
            "---",
            "",
            "*Vendor-neutral – Sortierung alphabetisch – Farben nach Okabe-Ito-Palette*",
            "",
            "_Erstellt von CHIMERA Suite v1.0.0_",
        ]

        path.write_text("\n".join(lines), encoding="utf-8")
        return path

    # ------------------------------------------------------------------
    # XLSX (openpyxl)
    # ------------------------------------------------------------------

    def _write_xlsx(
        self,
        rows: List[Dict[str, Any]],
        base_name: str,
    ) -> Optional[pathlib.Path]:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            import warnings
            warnings.warn(
                "openpyxl nicht installiert – XLSX-Export übersprungen. "
                "Installieren mit: pip install openpyxl",
                stacklevel=3,
            )
            return None

        path = self._out / f"{base_name}.xlsx"
        wb = openpyxl.Workbook()

        # --- Blatt 1: Zusammenfassung ---
        ws_sum = wb.active
        ws_sum.title = "Zusammenfassung"
        agg = _aggregate_by_workload(rows)

        header_fill = PatternFill("solid", fgColor="0173B2")
        header_font = Font(bold=True, color="FFFFFF")
        headers = [
            "Workload", "Throughput (ops/s)", "p50 ms", "p95 ms", "p99 ms",
            "Fehler (∅)", "Samples",
        ]
        ws_sum.append(headers)
        for cell in ws_sum[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for r in agg:
            tput = r.get("throughput_ops_per_sec_mean", r.get("throughput_ops_per_sec", 0))
            p50  = r.get("p50_latency_ms_mean",  r.get("p50_latency_ms", 0))
            p95  = r.get("p95_latency_ms_mean",  r.get("p95_latency_ms", 0))
            p99  = r.get("p99_latency_ms_mean",  r.get("p99_latency_ms", 0))
            errs = r.get("error_count_mean",      r.get("error_count", 0))
            n    = r.get("samples", 1)
            ws_sum.append([r["workload_id"], tput, p50, p95, p99, errs, n])

        # Spaltenbreiten
        for col, width in zip("ABCDEFG", [22, 20, 10, 10, 10, 12, 10]):
            ws_sum.column_dimensions[col].width = width

        # --- Blatt 2: Rohdaten ---
        ws_raw = wb.create_sheet("Rohdaten")
        if rows:
            all_keys = list(dict.fromkeys(k for r in rows for k in r))
            ws_raw.append(all_keys)
            for cell in ws_raw[1]:
                cell.font = Font(bold=True)
            for r in rows:
                ws_raw.append([r.get(k, "") for k in all_keys])

        # --- Blatt 3: Metadaten ---
        ws_meta = wb.create_sheet("Metadaten")
        ws_meta.append(["Schlüssel", "Wert"])
        ws_meta.append(["Erstellt am", self._generated_at])
        ws_meta.append(["CHIMERA Version", "1.0.0"])
        ws_meta.append(["Datensätze gesamt", len(rows)])
        ws_meta.append(["Neutralitätszertifiziert", "Ja"])

        wb.save(path)
        return path


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def _fmt(value: Any) -> str:
    """Formatiert einen Zahlenwert als String mit 2 Nachkommastellen."""
    if isinstance(value, float):
        return f"{value:.2f}"
    return str(value) if value is not None else "–"
